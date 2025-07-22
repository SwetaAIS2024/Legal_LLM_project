import openpyxl
from pathlib import Path

def read_list_from_file(file_path: str, delimiter: str = 'SPRTION') -> list:
    """Read file content and split it by a delimiter into a list, removing empty entries."""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read().strip()
    return [item.strip() for item in content.split(delimiter) if item.strip()]

def update_excel_with_data(excel_path: str, new_excel_path: str, clues_list: list, ans_list: list):
    """Open an Excel workbook, update rows where column B contains a question mark,
    writing clues into column D and answers into column C, then save to a new file."""
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    clue_idx = 0
    ans_idx = 0

    for row in range(2, sheet.max_row + 1):
        cell_b = sheet.cell(row=row, column=2)
        if cell_b.value and '?' in str(cell_b.value):
            if clue_idx < len(clues_list) and ans_idx < len(ans_list):
                sheet.cell(row=row, column=4, value=clues_list[clue_idx])
                sheet.cell(row=row, column=3, value=ans_list[ans_idx])
                clue_idx += 1
                ans_idx += 1
            else:
                break

    workbook.save(new_excel_path)
    print(f"Updated Excel saved to: {new_excel_path}")

def main():
    clues_file = r'c:/users/admin/desktop/tender/clues.TXT'
    combined_ans_file = r'c:/users/admin/desktop/tender/combined_ans.TXT'
    excel_file_path = r'C:\Users\admin\Desktop\tender\final_output.xlsx'
    new_excel_file_path = r'C:\Users\admin\Desktop\tender\final_output_modified.xlsx'
    
    clues_list = read_list_from_file(clues_file, delimiter='SPRTION')
    ans_list = read_list_from_file(combined_ans_file, delimiter='SPRTION')
    
    update_excel_with_data(excel_file_path, new_excel_file_path, clues_list, ans_list)

if __name__ == '__main__':
    main()
